import openpyxl

path="D:\\Azhar Python\\test1.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
sheet.cell(row=2,column=4).value="welcome"
#for r in range(1,3):
#    for c in range(1,3):
#        sheet.cell(row=r,column=c).value="welcome"
workbook.save(path)